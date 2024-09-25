import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# 用于对比各个文件的字典页的一桶
# 可以一次对比多个excel文档，并在每个文件中标黄（该文件和其他所有文件的差异）
# 可以指定sheet页名称（只能指定一个名称，所有文件的sheet页都应该相同）
# 会生成日志文件，日志文件观看方法如下：
# Comparing 【基地审核】设备编码清单（石家庄-环境监测系统）.xlsx with 【基地审核】设备编码清单（石家庄-压缩空气系统）.xlsx
# Difference in cell (9,4) of 【基地审核】设备编码清单（石家庄-环境监测系统）.xlsx compared to 【基地审核】设备编码清单（石家庄-压缩空气系统）.xlsx
# 说明：第2行表示：文件1和文件2对比，第2行表示：在第9行第4列这个单元格存在差异
# 说明：每处差异都有一行
# 注意：如果在文件2中，多插入了一行，那么该行及下面的所有行的单元格，都会标黄

# 当前日期和时间
now = datetime.now()

# 用户输入部分
# 定义需要对比的Excel文件列表
files = [
    r'C:\Users\JA085914\Desktop\PY\多文件对比\【基地审核】设备编码清单（石家庄-环境监测系统）.xlsx',
    r'C:\Users\JA085914\Desktop\PY\多文件对比\【基地审核】设备编码清单（石家庄-冷却水系统）.xlsx',
    r'C:\Users\JA085914\Desktop\PY\多文件对比\【基地审核】设备编码清单（石家庄-暖通空调系统）.xlsx',
    r'C:\Users\JA085914\Desktop\PY\多文件对比\【基地审核】设备编码清单（石家庄-天然气系统）.xlsx',
    r'C:\Users\JA085914\Desktop\PY\多文件对比\【基地审核】设备编码清单（石家庄-压缩空气系统）.xlsx',
    r'C:\Users\JA085914\Desktop\PY\多文件对比\To-Internal_设备编码清单_污水处理_SJ.xlsx',
]

# 日志文件路径
log_file_path = r'C:\Users\JA085914\Desktop\PY\多文件对比'

# 用户指定的日志文件名（带时间戳）
log_filename = f'comparison_log_{now.strftime("%Y-%m-%d_%H%M%S")}.txt'

# 完整的日志文件路径
log_full_path = os.path.join(log_file_path, log_filename)

# 定义需要对比的tab页名称
sheet_name_to_compare = "字典-设备&仪表管理"

# 定义需要标黄的颜色
highlight_color = 'FFFF00'  # 标黄的颜色

# 定义标黄的样式
yellow_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type='solid')

# 创建一个日志文件来记录差异
log_file = open(log_full_path, 'w')
log_file.write(f"Comparison log of tab '{sheet_name_to_compare}'\n\n")

# 清除指定 sheet 页中的所有单元格填充颜色
def clear_cell_fills(workbook, sheet_name):
    ws = workbook[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(fill_type=None)

# 遍历每个文件作为基准文件
for base_file_path in files:
    base_wb = load_workbook(base_file_path)

    # 清除指定 sheet 页中的所有单元格填充颜色
    if sheet_name_to_compare in base_wb.sheetnames:
        clear_cell_fills(base_wb, sheet_name_to_compare)
    else:
        log_file.write(f"Sheet '{sheet_name_to_compare}' not found in {os.path.basename(base_file_path)}\n")
        continue

    # 读取基准文件的所有工作表
    xls_base = pd.ExcelFile(base_file_path)

    # 对其他每个文件进行比较
    for compare_file_path in files:
        if base_file_path != compare_file_path:
            print(f"Comparing {os.path.basename(base_file_path)} with {os.path.basename(compare_file_path)}")
            log_file.write(f"\nComparing {os.path.basename(base_file_path)} with {os.path.basename(compare_file_path)}\n")

            # 读取比较文件的所有工作表
            xls_compare = pd.ExcelFile(compare_file_path)

            # 检查tab页是否存在
            if sheet_name_to_compare in xls_base.sheet_names and sheet_name_to_compare in xls_compare.sheet_names:
                df_base = pd.read_excel(base_file_path, sheet_name=sheet_name_to_compare)
                df_compare = pd.read_excel(compare_file_path, sheet_name=sheet_name_to_compare)

                # 确保两个DataFrame的索引和列标签一致
                df_base, df_compare = df_base.align(df_compare, axis=0)
                df_base, df_compare = df_base.align(df_compare, axis=1)

                # 打印对齐后的DataFrame以检查
                print(f"Base DataFrame after alignment:")
                print(df_base.head())
                print(f"Compare DataFrame after alignment:")
                print(df_compare.head())

                # 对比两个DataFrame，找出不同的单元格
                diff_mask = (df_base != df_compare) & ~(df_base.isnull() & df_compare.isnull())

                # 获取工作表
                ws_base = base_wb[sheet_name_to_compare]

                # 遍历不同的单元格，并将其标黄
                for r, row in enumerate(diff_mask.values, start=2):  # 从第2行开始（Excel中的第1行）
                    for c, value in enumerate(row, start=1):
                        if value:
                            cell = ws_base.cell(row=r, column=c)
                            cell.fill = yellow_fill
                            # 记录差异
                            log_file.write(f"Difference in cell ({r},{c}) of {os.path.basename(base_file_path)} compared to {os.path.basename(compare_file_path)}\n")
                            # 打印具体差异值
                            base_value = df_base.iat[r-2, c-1]  # 转换为Pandas索引
                            compare_value = df_compare.iat[r-2, c-1]  # 转换为Pandas索引
                            print(f"Base value: {base_value}, Compare value: {compare_value}")
            else:
                log_file.write(f"Sheet '{sheet_name_to_compare}' not found in one or more files.\n")

    # 保存修改后的Workbook
    base_wb.save(base_file_path)

# 关闭日志文件
log_file.close()