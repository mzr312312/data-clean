import openpyxl
from copy import copy

# 定义文件路径
template_path = r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\0 梳理工具\数据点编码的模板.xlsx'
target_files = [
    r'C:\Users\JA085914\Desktop\实验\To-Internal_设备类型清单&数据点清单_暖通空调_SJ.xlsx',
]

# 打开模板文件
template_wb = openpyxl.load_workbook(template_path)
template_ws = template_wb['信号字典']

for file in target_files:
    # 加载目标文件
    target_wb = openpyxl.load_workbook(file)

    # 如果目标文件中存在“信号字典”sheet，则删除它
    if '信号字典' in target_wb.sheetnames:
        del target_wb['信号字典']

    # 创建新的“信号字典”sheet
    target_ws = target_wb.create_sheet('信号字典')

    # 复制模板中的“信号字典”sheet内容和格式到目标文件
    for row in template_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # 保存目标文件
    target_wb.save(file)
