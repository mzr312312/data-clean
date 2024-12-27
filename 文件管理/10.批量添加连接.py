import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook

# 读取Excel文件
excel_path = r'I:\抖音收藏\文件列表.xlsx'
df = pd.read_excel(excel_path)

# 获取“原始文件名”列
original_filenames = df['原始文件名'].tolist()

# 定义文件夹路径
base_folder = Path(r'I:\抖音收藏')

# 创建一个字典来存储文件名到绝对路径的映射
file_path_map = {}

# 遍历文件夹及其子文件夹
for root, dirs, files in os.walk(base_folder):
    for file in files:
        if file in original_filenames:
            # 获取文件的绝对路径
            absolute_path = Path(root) / file
            file_path_map[file] = str(absolute_path)

# 在新列中添加绝对路径
df['链接'] = df['原始文件名'].map(file_path_map)

# 将结果保存回Excel文件
df.to_excel(excel_path, index=False)

# 使用openpyxl插入超链接
wb = load_workbook(excel_path)
ws = wb.active

# 遍历新添加的“链接”列，插入超链接
for row in range(2, ws.max_row + 1):
    link_value = ws.cell(row=row, column=3).value  # 假设“链接”列是第三列
    if link_value:
        # 插入超链接，指向视频文件的绝对路径
        ws.cell(row=row, column=3).hyperlink = f"file:///{link_value}"
        ws.cell(row=row, column=3).value = f"打开 {ws.cell(row=row, column=1).value}"
        ws.cell(row=row, column=3).style = 'Hyperlink'

# 保存修改后的Excel文件
wb.save(excel_path)

print("处理完成，结果已保存到原Excel文件中。")