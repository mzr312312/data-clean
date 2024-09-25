
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import MultiCellRange


import xlwings as xw

# 批量同步的是”['字典-设备&仪表管理']“的表格
# 针对的是，设备编码清单
# 模板文件路径
template_path = r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\0 梳理工具\字典模板-SJ.xlsx'
# 目标文件列表
target_files = [
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\15.纯水系统\To-Base_设备编码清单_纯水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\04.大宗气体系统\To-Base_设备编码清单_大宗气体_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\16.工艺冷却水系统\To-Base_设备编码清单_工艺冷却水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\06.化学品系统\To-Base_设备编码清单_化学品_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\10.环境监测系统\To-Base_设备编码清单_环境监测_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\02.暖通空调系统\To-Base_设备编码清单_暖通空调_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\09.排气系统\To-Base_设备编码清单_排气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\21.生产系统-电池\To-Base_设备编码清单_生产-电池_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\20.生产系统-切片\To-Base_设备编码清单_生产-切片_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\14.生活用水系统\To-Base_设备编码清单_生活用水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\00 待定 水系统\To-Base_设备编码清单_水系统_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\05.特气系统\To-Base_设备编码清单_特气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\07.天然气系统\To-Base_设备编码清单_天然气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\11.污水处理系统\To-Base_设备编码清单_污水处理_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\03.压缩空气系统\To-Base_设备编码清单_压缩空气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\17.中水系统\To-Base_设备编码清单_中水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\15.纯水系统\To-Internal_设备编码清单_纯水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\04.大宗气体系统\To-Internal_设备编码清单_大宗气体_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\16.工艺冷却水系统\To-Internal_设备编码清单_工艺冷却水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\06.化学品系统\To-Internal_设备编码清单_化学品_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\10.环境监测系统\To-Internal_设备编码清单_环境监测_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\02.暖通空调系统\To-Internal_设备编码清单_暖通空调_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\09.排气系统\To-Internal_设备编码清单_排气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\21.生产系统-电池\To-Internal_设备编码清单_生产-电池_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\20.生产系统-切片\To-Internal_设备编码清单_生产-切片_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\14.生活用水系统\To-Internal_设备编码清单_生活用水_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\00 待定 水系统\To-Internal_设备编码清单_水系统_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\05.特气系统\To-Internal_设备编码清单_特气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\07.天然气系统\To-Internal_设备编码清单_天然气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\11.污水处理系统\To-Internal_设备编码清单_污水处理_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\03.压缩空气系统\To-Internal_设备编码清单_压缩空气_SJ.xlsx',
    r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\17.中水系统\To-Internal_设备编码清单_中水_SJ.xlsx',

]

# 打开模板文件
try:
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb['字典-设备&仪表管理']
    print(f"成功打开模板文件: {template_path}")
except Exception as e:
    print(f"无法打开模板文件: {template_path}, 错误信息: {e}")
    exit()

# 定义复制函数（与之前相同）
def copy_border(src_cell, target_cell):
    if src_cell.border.top.style is not None:
        target_cell.border = Border(
            top=Side(style=src_cell.border.top.style, color=src_cell.border.top.color),
            bottom=Side(style=src_cell.border.bottom.style, color=src_cell.border.bottom.color),
            left=Side(style=src_cell.border.left.style, color=src_cell.border.left.color),
            right=Side(style=src_cell.border.right.style, color=src_cell.border.right.color)
        )

def copy_font(src_cell, target_cell):
    color_rgb = src_cell.font.color.rgb if isinstance(src_cell.font.color.rgb, str) else '000000'
    target_cell.font = Font(
        name=src_cell.font.name,
        size=src_cell.font.size,
        bold=src_cell.font.bold,
        italic=src_cell.font.italic,
        vertAlign=src_cell.font.vertAlign,
        underline=src_cell.font.underline,
        strike=src_cell.font.strike,
        color=Color(rgb=color_rgb)
    )

def copy_alignment(src_cell, target_cell):
    target_cell.alignment = Alignment(
        horizontal=src_cell.alignment.horizontal,
        vertical=src_cell.alignment.vertical,
        text_rotation=src_cell.alignment.text_rotation,
        wrap_text=src_cell.alignment.wrap_text,
        shrink_to_fit=src_cell.alignment.shrink_to_fit,
        indent=src_cell.alignment.indent
    )

def copy_fill(src_cell, target_cell):
    target_cell.fill = PatternFill(
        fill_type=src_cell.fill.fill_type,
        start_color=src_cell.fill.start_color,
        end_color=src_cell.fill.end_color
    )

def copy_column_dimensions(src_ws, target_ws):
    for col in src_ws.column_dimensions:
        target_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width

def copy_data_validations(src_ws, target_ws):
    for dv in src_ws.data_validations.dataValidation:
        new_dv = DataValidation(
            type=dv.type, formula1=dv.formula1, formula2=dv.formula2,
            allow_blank=dv.allow_blank, showErrorMessage=dv.showErrorMessage,
            errorTitle=dv.errorTitle, error=dv.error,
            showInputMessage=dv.showInputMessage, promptTitle=dv.promptTitle, prompt=dv.prompt
        )
        if isinstance(dv.sqref, str):
            new_dv.sqref = MultiCellRange(dv.sqref)
        else:
            new_dv.sqref = dv.sqref
        target_ws.add_data_validation(new_dv)



# 打开目标 Excel 文件
for file in target_files:
    try:
        # 启动 Excel 应用程序（在后台运行）
        app = xw.App(visible=False)

        # 打开目标文件
        wb = app.books.open(file)
        print(f"成功打开目标文件: {file}")

        # 检查是否存在 '字典-设备&仪表管理' sheet，如果存在则删除
        if '字典-设备&仪表管理' in [sheet.name for sheet in wb.sheets]:
            wb.sheets['字典-设备&仪表管理'].delete()
            print(f"删除目标文件中的 '字典-设备&仪表管理' sheet: {file}")

        # 创建新的 '字典-设备&仪表管理' sheet 并复制模板中的内容
        new_sheet = wb.sheets.add('字典-设备&仪表管理', before=wb.sheets[0])
        template_wb = xw.Book(template_path)
        template_sheet = template_wb.sheets['字典-设备&仪表管理']

        # 复制模板 sheet 的所有内容
        template_sheet.range('A1').expand().copy(new_sheet.range('A1'))
        print(f"复制模板中的内容到目标文件: {file}")

        # 保存并关闭文件
        wb.save(file)
        wb.close()
        print(f"成功保存目标文件: {file}")

        # 关闭模板文件
        template_wb.close()

    except Exception as e:
        print(f"处理文件 {file} 时出错，错误信息: {e}")

    finally:
        app.quit()

print("所有文件已更新完毕。")
