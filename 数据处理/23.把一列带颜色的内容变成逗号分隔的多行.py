import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# 定义输入文件和输出文件路径
input_file = r"..\PY\数据处理\23.把一列带颜色的内容变成逗号分隔的多行.xlsx"
output_dir = r"..\PY\数据处理"

# 加载工作簿和工作表
wb = load_workbook(input_file)
ws = wb.active

# 获取第一列的数据和填充色
data = []
colors = []

for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):  # 假设第一行为标题行
    cell = row[0]
    data.append(cell.value)
    if cell.fill.fgColor.type == "rgb":
        colors.append(cell.fill.fgColor.rgb)
    elif cell.fill.fgColor.type == "indexed":
        colors.append(cell.fill.fgColor.indexed)
    else:
        colors.append(None)

# 合并同类单元格
merged_data = []
current_group = []
current_color = None

for i, (value, color) in enumerate(zip(data, colors)):
    if color == current_color:
        current_group.append(value)
    else:
        if current_group:
            merged_data.append(",".join(str(v) for v in current_group))
        current_group = [value]
        current_color = color

# 添加最后一组数据
if current_group:
    merged_data.append(",".join(str(v) for v in current_group))

# 创建新的工作簿并写入合并后的数据
new_wb = Workbook()  # 创建新的工作簿
new_ws = new_wb.active
new_ws.title = "合并结果"

# 写入标题
new_ws.cell(row=1, column=1, value="合并后内容")

# 写入合并后的数据
for i, value in enumerate(merged_data, start=2):
    new_ws.cell(row=i, column=1, value=value)

# 生成带时间戳的输出文件名
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
output_file = os.path.join(output_dir, f"合并列供导入_{timestamp}.xlsx")

# 保存输出文件
new_wb.save(output_file)

print(f"处理完成，输出文件已保存为：{output_file}")