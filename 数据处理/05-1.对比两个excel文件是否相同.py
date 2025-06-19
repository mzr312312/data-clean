import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
# 本功能需要两个excel文件具有相同的标sheet名称，且具有相同的行索引和列标签，且具有相同的列标签和索引标签
# 本功能会生成一个新文件，在新文件中将部分单元格标黄，以便对比两个excel文件是否相同
# 可以选在在哪个工作簿上进行标黄，当前是默认在file2上进行标黄（一般来说，file1上是旧版本，file2上是最新版本）

# 自定义变量
# 定义需要对比的两个Excel文件路径
file1_path = r'..\PY\数据处理\11.根据已有成果，生成标准字典（特征字段-数据点映射）\标准字典（特征字段-数据点映射）.xlsx'
file2_path = r'..\PY\数据处理\11.根据已有成果，生成标准字典（特征字段-数据点映射）\标准字典（特征字段-数据点映射） - 副本.xlsx'
# 定义需要标黄的颜色
highlight_color = 'FFFF00'  # 标黄的颜色
# 定义需要在哪个工作簿上进行标黄（二选一）
highlight_workbook = 'file2'  # 选择在哪个工作簿上进行标黄，可选值：'file1' 或 'file2'

# 定义新生成的文件存放的位置
# 获取当前日期和时间
now = datetime.now()
output_file_name = f'【对比结果】_{now.strftime("%Y-%m-%d_%H%M")}.xlsx'
output_file_path = rf'C:\Users\JA085914\Desktop\PY\多文件对比\{output_file_name}'


# 以下内容不需要修改
# 读取两个Excel文件
file1 = file1_path
file2 = file2_path

# 读取所有工作表
xls1 = pd.ExcelFile(file1)
xls2 = pd.ExcelFile(file2)

# 加载需要标黄的Excel文件的Workbook
if highlight_workbook == 'file1':
    wb = load_workbook(file1)
else:
    wb = load_workbook(file2)

# 定义标黄的样式
yellow_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type='solid')

# 遍历所有工作表
for sheet_name in xls1.sheet_names:
    if sheet_name in xls2.sheet_names:
        df1 = pd.read_excel(file1, sheet_name=sheet_name)
        df2 = pd.read_excel(file2, sheet_name=sheet_name)

        # 确保两个DataFrame的索引和列标签一致
        df1, df2 = df1.align(df2, axis=0)
        df1, df2 = df1.align(df2, axis=1)

        # 对比两个DataFrame，找出不同的单元格
        diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())

        # 获取工作表
        ws = wb[sheet_name]

        # 遍历不同的单元格，并将其标黄
        for r, row in enumerate(diff_mask.values, start=2):  # 从第2行开始，即整体下移一行
            for c, value in enumerate(row, start=1):
                if value:
                    cell = ws.cell(row=r, column=c)
                    cell.fill = yellow_fill

# 保存修改后的Workbook
wb.save(output_file_path)
