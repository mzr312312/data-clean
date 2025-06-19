import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import range_boundaries
from datetime import datetime

# 定义输入文件路径和输出文件路径（相对路径）
input_file = r"../PY/数据处理/24.合并单元格拆开后向下填充.xlsx"
output_dir = os.path.dirname(input_file)

# 读取 Excel 文件
df = pd.read_excel(input_file)

# 创建一个新的 Excel 工作簿
output_timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
output_file = os.path.join(output_dir, f"合并单元格{output_timestamp}.xlsx")

# 使用 openpyxl 创建一个新的工作簿并写入数据
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 将 DataFrame 写入工作表
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 遍历每一列，找到需要合并的单元格范围
for col_idx, col in enumerate(df.columns, start=1):  # 列索引从 1 开始
    start_row = None
    for row_idx, value in enumerate(df[col], start=2):  # 行索引从 2 开始（跳过标题行）
        if pd.notna(value):  # 如果当前单元格有值
            if start_row is not None:  # 如果之前有需要合并的范围
                ws.merge_cells(start_row=start_row, start_column=col_idx,
                               end_row=row_idx - 1, end_column=col_idx)
            start_row = row_idx  # 更新起始行
    # 处理最后一段需要合并的范围
    if start_row is not None and start_row < len(df) + 1:
        ws.merge_cells(start_row=start_row, start_column=col_idx,
                       end_row=len(df) + 1, end_column=col_idx)

# 保存工作簿
wb.save(output_file)

print(f"合并单元格完成，输出文件已保存为：{output_file}")