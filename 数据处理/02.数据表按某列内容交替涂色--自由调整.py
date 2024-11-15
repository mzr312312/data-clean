import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from datetime import datetime
# 没有选择sheet页的功能，最好新建一个excel文档，把需要涂色的sheet页拷贝过去
# 可配置变量
# Excel文件路径--现有未涂色的文件
file_path = r'../../PY/数据处理/02.数据表按某列内容交替涂色--自由调整.xlsx'
# 新生成的Excel文件路径--新生成的文件
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # 生成时间戳，格式为：年月日_时分秒
new_file_path = fr'../../PY/数据处理/涂色文件_{timestamp}.xlsx'
# 设置颜色--用Quicker取色即可
# 随便设置一个颜色就行，后面还要用excel，按某一列的颜色排序后，把所有颜色都改成正经颜色
start_color = 'FFEBF1DE'  # 绿色
end_color = 'FFEBF1DE'  # 绿色
# 涂色依据的列名称--希望依据哪一列的内容涂色，只需要把列的名称改成和Excel文件相同即可
color_by_column = '设备/仪表原有编号'

# 读取Excel文件
df = pd.read_excel(file_path)

# 创建新的工作簿
wb = Workbook()
ws = wb.active

# 定义涂色的填充样式
gray_fill = PatternFill(start_color=start_color, end_color=end_color, fill_type='solid')


# 确定需要涂色的系统名称
def should_color(system_name, prev_system_name, prev_should_color):
    # 当前系统名称与前一个不同，根据奇偶性决定是否涂色
    if system_name != prev_system_name:
        # 如果是第一次遇到系统名称，则默认涂色
        if prev_system_name is None:
            return True
        # 从第二个系统开始，涂色与前一个系统相反
        else:
            return not prev_should_color
    # 如果系统名称相同，则使用前一个系统的涂色状态
    else:
        return prev_should_color


# 将DataFrame写入新的工作表，并根据规则涂色
prev_system_name = None
prev_should_color = None
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

        # 确定是否需要涂色
        if r_idx > 0:  # 跳过标题行
            system_name = df.loc[r_idx - 1, color_by_column]
            cell_should_color = should_color(system_name, prev_system_name, prev_should_color)
            prev_system_name = system_name
            prev_should_color = cell_should_color

            if cell_should_color:
                cell.fill = gray_fill

# 保存新的文件
wb.save(new_file_path)