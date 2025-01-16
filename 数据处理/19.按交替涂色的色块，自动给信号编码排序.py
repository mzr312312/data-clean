import pandas as pd
from openpyxl import Workbook
from datetime import datetime
import openpyxl

"""

此脚本的逻辑是：
1.需要代码读取单元格颜色，只有“连续的，且颜色相同的单元格”才能成为“同一个色块“
2.在同一个色块下，进行当前逻辑的编号，如果遍历到了一个新的色块，则一定从1开始，重新编号
3.即使新色块的第一行和上一个色块的最后一行的内容是一样的，但是他们的颜色不一样，他们就属于不同的色块，不同的色块，编号不连续

使用说明：
该脚本用于处理Excel文件中的一列数据，依据数据的色块进行编码，并生成一个新的Excel文件。具体功能如下：
1. 脚本会读取指定路径的Excel文件，并识别目标列中的数据，按照色块进行分组。
2. 在每个色块内部，脚本会对数据进行编号：
   - 重复的数据会依次编号，从1开始，比如：1, 2, 3, 4
   - 不重复的数据则会从1重新开始编号。
3. 用户需注意，源文件中填写的数据列必须保持列名不变，以便系统能够正确识别。
4. 编码后的结果将会写入新的Excel文件中，新文件将以时间戳命名，并保存在指定路径。

使用步骤：
1. 在Excel的指定列中填写数据，并确保数据按设备类型分类并带有色块。
2. 调整源文件列中数据的顺序，确保不破坏色块的结构。
3. 运行该脚本，脚本将自动对数据进行编码并生成新的Excel文件。

注意事项：
- 确保安装了‘pandas’和‘openpyxl’库，以便正常执行文件读写操作。
- 本脚本适用于按照色块分组的数据编码，不适合其他格式的数据处理。
"""


# 文件路径
file_path = r'../../PY/数据处理/19.按交替涂色的色块，自动给信号编码排序.xlsx'
# 新生成的文件路径
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
new_file_path = fr'../../PY/数据处理/数据点编号_{timestamp}.xlsx'

# 设置连续空值的最大数量
max_empty_count = 20000  # 你可以根据需要修改这个值

# 读取Excel文件
wb = openpyxl.load_workbook(file_path)
sheet = wb.active  # 获取活动的工作表

# 涉及的列
target_column_index = 1  # 根据你的目标列修改，如第一列为1，第二列则为2
previous_value = None
previous_color = None
current_block = []
results = []
empty_count = 0  # 计数器，用于跟踪连续空值的数量

# 遍历目标列，识别色块
for row in sheet.iter_rows(min_row=2, min_col=target_column_index, max_col=target_column_index):  # 从第二行开始
    cell = row[0]
    value = cell.value
    fill_color = cell.fill.start_color.index  # 获取单元格颜色

    # 如果值为空，增加空值计数器
    if value is None:
        empty_count += 1
        # 如果连续空值数量达到最大值，停止处理
        if empty_count >= max_empty_count:
            print(f"连续遇到 {max_empty_count} 个空值，停止处理。")
            break
        continue
    else:
        empty_count = 0  # 重置空值计数器

    # 判断当前单元格的值和颜色是否与上一个相同
    if value == previous_value and fill_color == previous_color:
        current_block.append(value)  # 如果相同，继续添加到当前色块
    else:
        # 当前值与前一个值不同，处理当前色块
        if current_block:
            # 为当前色块编号
            count_map = {}
            for item in current_block:
                if item in count_map:
                    count_map[item] += 1
                else:
                    count_map[item] = 1
                results.append((item, count_map[item]))

        # 开启新的色块
        current_block = [value]

    # 更新前一个值和前一个颜色
    previous_value = value
    previous_color = fill_color

# 处理最后一个色块
if current_block:
    count_map = {}
    for item in current_block:
        if item in count_map:
            count_map[item] += 1
        else:
            count_map[item] = 1
        results.append((item, count_map[item]))

# 创建新的工作簿
wb_output = Workbook()
ws = wb_output.active

# 写入标题
ws.append(['原始内容', '编号'])

# 写入结果
for original_value, number in results:
    ws.append([original_value, number])

# 保存新的文件
wb_output.save(new_file_path)

print(f"新的文件已生成：{new_file_path}")