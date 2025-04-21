import os
from openpyxl import load_workbook
from datetime import datetime

"""
# 脚本功能说明：
# 本脚本用于处理 Excel 文件中的多列数据，按照以下规则合并单元格：
# 1. 连续单元格的内容相同且填充颜色一致时，将其合并为一个单元格。
# 2. 输入文件包含 12 列，列名分别为“第1列”、“第2列”、“第3列”……“第12列”。
# 3. 脚本会自动识别这些列，并对每一列执行合并操作。

# 使用方法：
# 1. 确保输入文件路径正确：
#    - 输入文件路径为相对路径 "../../PY/数据处理/22.把一列内容里相同内容合并单元格.xlsx"。
#    - 如果文件路径不同，请修改 `input_file` 的值。
# 2. 确保输入文件格式符合要求：
#    - 文件必须是标准的 .xlsx 格式。
#    - 第一行必须包含列名，列名格式为“第X列”（例如“第1列”、“第2列”）。
#    - 数据从第二行开始。
# 3. 运行脚本后，输出文件将保存在与输入文件相同的目录下，文件名为“合并单元格_时间戳.xlsx”。

# 注意事项：
# 1. 单元格填充颜色必须通过 Excel 的标准功能设置，否则可能无法正确读取颜色。
# 2. 如果某些列没有满足合并条件的数据，则不会对该列进行任何操作。
# 3. 输出文件中，只有内容和填充颜色都相同的连续单元格会被合并。

# 示例输入文件结构：
# | 第1列 | 第2列 | 第3列 |
# |-------|-------|-------|
# | 1     | A     | X     |
# | 1     | A     | Y     |
# | 1     | B     | Z     |
# | 2     | C     | Z     |
# | 2     | C     | Z     |

# 示例输出文件结构：
# | 第1列 | 第2列 | 第3列 |
# |-------|-------|-------|
# | 1     | A     | X     |
# |       |       | Y     |
# |       | B     | Z     |
# | 2     | C     |       |
# |       |       |       |
"""

# 获取脚本所在目录
script_dir = os.path.dirname(os.path.abspath(__file__))

# 使用你提供的相对路径（原始字符串）
input_file = os.path.join(script_dir, r"..\..\PY\数据处理", "22.把一列内容里相同内容合并单元格.xlsx")

# 输出调试信息
print(f"输入文件路径：{input_file}")

# 检查文件是否存在
if not os.path.exists(input_file):
    print(f"文件不存在：{input_file}")
    exit()

# 加载 Excel 文件
try:
    workbook = load_workbook(input_file)
    sheet = workbook["Sheet1"]  # 假设工作表名称为 "Sheet1"
except Exception as e:
    print(f"无法加载文件或工作表：{e}")
    exit()

# 获取标题行（假设标题行在第1行）
header_row = 1
columns_to_process = []

# 动态获取列名（“第1列”到“第12列”）
for col in range(1, sheet.max_column + 1):
    header_value = sheet.cell(row=header_row, column=col).value
    if header_value and header_value.startswith("第") and header_value.endswith("列"):
        columns_to_process.append(col)

if not columns_to_process:
    print("未找到符合命名规则的列（‘第X列’）")
    exit()

print(f"需要处理的列：{columns_to_process}")

# 定义合并逻辑函数
def merge_cells_based_on_content_and_fill(sheet, target_column):
    start_row = header_row + 1  # 数据从标题行的下一行开始
    end_row = sheet.max_row

    current_value = None
    current_fill = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=target_column)
        cell_value = cell.value
        cell_fill = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else None  # 获取填充颜色的 RGB 值

        # 打印调试信息
        print(f"列 {target_column}, 行 {row}: 内容={cell_value}, 填充颜色={cell_fill}")

        # 如果当前值或填充颜色与上一个不同，检查是否需要合并
        if cell_value != current_value or cell_fill != current_fill:
            # 如果有需要合并的单元格，执行合并
            if merge_start is not None and merge_start < row - 1:
                sheet.merge_cells(
                    start_row=merge_start,
                    start_column=target_column,
                    end_row=row - 1,
                    end_column=target_column
                )
                print(f"合并单元格：列 {target_column}, 行 {merge_start} 到 行 {row - 1}")
            # 更新当前值、填充颜色和合并起始行
            current_value = cell_value
            current_fill = cell_fill
            merge_start = row

    # 处理最后一组连续相同内容和颜色
    if merge_start is not None and merge_start < end_row:
        sheet.merge_cells(
            start_row=merge_start,
            start_column=target_column,
            end_row=end_row,
            end_column=target_column
        )
        print(f"合并单元格：列 {target_column}, 行 {merge_start} 到 行 {end_row}")

# 遍历所有需要处理的列，依次执行合并逻辑
for col in columns_to_process:
    print(f"正在处理列：{col}")
    merge_cells_based_on_content_and_fill(sheet, col)

# 保存结果文件
output_dir = os.path.dirname(input_file)
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
output_file = os.path.join(output_dir, f"合并单元格_{timestamp}.xlsx")
workbook.save(output_file)

print(f"处理完成，结果已保存到：{output_file}")