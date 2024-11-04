import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
# 此脚本可以根据某一列的交替颜色，统计每一个色块有多少行
# 因为颜色是按行交替排列的，定义每一段有连续颜色的行为一个“组”
# 需要统计每一个”组“有多少行
# 输出结果为一个新的Excel文件，列名为“数量统计”

# 自定义内容：
# 1. 输入文件路径
input_file_path = r"C:\Users\JA085914\Desktop\PY\数据处理\10.根据交替的颜色，数每一个色块有多少行.xlsx"

# 2. 输出文件路径--在同一个文件夹里自动生成，带有时间戳
output_dir = os.path.dirname(input_file_path)
output_filename = f"count_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
output_path = os.path.join(output_dir, output_filename)

# 3. 列名为“数量统计”--这是告诉程序，把统计后的数字写入哪一列
result_column = 2  # A列为1，B列为2，以此类推

# 4. 定义sheet名称，默认是“采集点”
sheet_name = '按色块数数量'

# 5. & 6. 定义以哪一列的涂色为准，去定义一个“组”；（定义每一段有连续颜色的行为一个“组”）
# 主要目的是看看各个组里，行数是不是一样的，从而判断他们是否可以作为同一个设备类型
color_check_column = 1  # E列为5

# 加载工作簿
wb = load_workbook(input_file_path)
ws = wb[sheet_name]

# 初始化变量
current_group_color = None
group_count = 0
group_counts = []

# 写入标题行
ws.cell(row=1, column=result_column).value = '数量统计'

# 遍历“设备类型”列，从第二行开始
for row in range(2, ws.max_row + 1):  # 从第二行开始
    cell = ws.cell(row=row, column=color_check_column)  # 用 column值来确定应该检查哪一列的颜色
    if cell.fill.start_color.index != '00000000':  # 检查是否有颜色填充
        if current_group_color is None or not current_group_color:
            # 开始新的有颜色的组
            current_group_color = True
            if group_count > 0:  # 如果是第一组以外的情况
                group_counts.append(group_count)
            group_count = 1
        else:
            group_count += 1
    else:
        if current_group_color:
            # 结束有颜色的组
            group_counts.append(group_count)
            group_count = 1
            current_group_color = False
        else:
            group_count += 1

# 添加最后一个组的计数
group_counts.append(group_count)

# 写入结果
current_group_index = 0
for row in range(2, ws.max_row + 1):  # 从第二行开始
    if (row == 2) or (ws.cell(row=row, column=color_check_column).fill.start_color.index != ws.cell(row=row - 1, column=color_check_column).fill.start_color.index):
        # 检查是否还有剩余的组计数
        if current_group_index < len(group_counts):
            ws.cell(row=row, column=result_column).value = group_counts[current_group_index]  # 将数量统计写入A列
            current_group_index += 1
        else:
            break  # 如果所有组计数都已经写入，则提前结束循环

# 保存到新文件
wb.save(output_path)

print(f"文件已保存至: {output_path}")