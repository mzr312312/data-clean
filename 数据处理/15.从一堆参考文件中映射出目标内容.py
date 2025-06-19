import openpyxl
import os
from datetime import datetime
# 是vlookpu功能的拓展，可以实现从多个文件的多个sheet页的多个列的单元格中，查找匹配的单元格并返回其值。
# 详见md文件：15.从一堆参考文件中映射出目标内容.md

# 数据源文件路径
data_source_path = r'..\PY\数据处理\电表全文映射\需要被映射的数据源.xlsx'

# 参考文件路径
reference_paths = [
    r'..\PY\数据处理\电表全文映射\新-电池车间电表清单.xlsx',
    r'..\PY\数据处理\电表全文映射\新-动力站电表清单.xlsx',
    r'..\PY\数据处理\电表全文映射\新-切片车间电表清单.xlsx'
]

# 创建一个字典来存储设备名称及其映射名称
mapping_dict = {}

# 读取数据源文件的设备名称
data_source_workbook = openpyxl.load_workbook(data_source_path)
data_source_sheet = data_source_workbook.active

# 按列名读取设备名称
header = [cell.value for cell in data_source_sheet[1]]  # 假设第一行为标题行
device_name_col_index = header.index('设备名称')  # 找到设备名称列的索引

device_names = [row[device_name_col_index].value for row in data_source_sheet.iter_rows(min_row=2) if row[device_name_col_index].value]

# 遍历参考文件
for reference_path in reference_paths:
    reference_workbook = openpyxl.load_workbook(reference_path)
    for sheet in reference_workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            # 遍历当前行的每一列
            for i in range(len(row)):
                if row[i] in device_names:
                    # 找到匹配的设备名称，记录右侧单元格的内容
                    if i + 1 < len(row):  # 确保右侧单元格存在
                        # 只有在映射字典中尚未存在该设备名称的情况下，才添加映射
                        if row[i] not in mapping_dict:
                            mapping_dict[row[i]] = row[i + 1]  # 将右侧单元格内容作为映射
                    # 继续检查这一行的其他列，不要立即跳出

# 生成新文件名称
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
result_file_path = os.path.join(os.path.dirname(data_source_path), f'ZZ结果_{timestamp}.xlsx')

# 创建结果文件
result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active
result_sheet.append(['数据源设备名称', '映射名称'])  # 添加标题

# 循环将设备名称和映射名称写入结果文件
for device in device_names:
    mapping_name = mapping_dict.get(device, None)
    result_sheet.append([device, mapping_name])

# 保存结果文件
result_workbook.save(result_file_path)
print(f'映射结果已保存到: {result_file_path}')
