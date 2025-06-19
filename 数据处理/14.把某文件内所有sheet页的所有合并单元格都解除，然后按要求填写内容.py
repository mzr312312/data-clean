import openpyxl
import os
from datetime import datetime

# 文件路径
file_path = r'..\PY\数据处理\14.把某文件内所有sheet页的所有合并单元格都解除，然后按要求填写内容.xlsx'
# 加载工作簿
workbook = openpyxl.load_workbook(file_path)

# 遍历所有工作表
for sheet in workbook.worksheets:
    # 获取合并单元格并转为列表
    merged_cells = list(sheet.merged_cells.ranges)
    for merged in merged_cells:
        # 获取合并单元格的内容
        value = sheet.cell(row=merged.min_row, column=merged.min_col).value
        # 解除合并
        sheet.unmerge_cells(str(merged))
        # 填充解除合并后产生的空单元格
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if sheet.cell(row=row, column=col).value is None:
                    sheet.cell(row=row, column=col).value = value

# 生成新文件名称
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
new_file_name = f'新生成文件_{timestamp}.xlsx'
new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)

# 保存新文件
workbook.save(new_file_path)
print(f'新文件已生成: {new_file_path}')
