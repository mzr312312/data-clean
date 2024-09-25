from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# 功能：
#   遍历表格某一列，如果涂色，则该行全部涂成指定颜色
# 步骤：
#   1）定义颜色（新表格涂成什么色）；
#   2)指定要检查的列号，如果这一列的某一行有背景色，则该行全部涂色；
#       如果column_to_check为3，那么cell_to_check将是row[2]（2代表B列）
#       举例：如果想检查D列，则需要cell_to_check=5
#   3）指定读取的源文件和生成的目标文件
#   4）指定读取的源文件中，工作表（tab页）的名称（因为源文件可能包含多个tab页）

# 定义颜色
color_index = 'FFFFF2CC'
fill = PatternFill(start_color=color_index,
                   end_color=color_index,
                   fill_type='solid')

# 指定要检查的列号
column_to_check = 17  # column_to_check = 5的话，是检查第4列

# 定义源文件路径
source_file_path = r'C:\Users\JA085914\Desktop\临时\PY实验\工作簿.xlsx'
# 定义目标文件路径
target_file_path = r'C:\Users\JA085914\Desktop\临时\PY实验\涂色_test.xlsx'

# 指定工作表名称
worksheet_name = '采集点'  # 假设工作表名为'Sheet1'

# 加载源Excel文件
workbook = load_workbook(filename=source_file_path)
# 获取指定的工作表
sheet = workbook[worksheet_name]

# 遍历每一行
for row in sheet.iter_rows(min_row=2, values_only=False):  # 从第二行开始遍历，假设第一行为标题行
    # 获取指定列的单元格
    cell_to_check = row[column_to_check - 1]  # 减1是因为列表索引是从0开始的

    if cell_to_check.fill.start_color.index != '00000000':  # 检查指定列的单元格是否有颜色填充
        # 给该行的所有单元格应用填充样式
        for cell in row:
            cell.fill = fill

# 保存到新文件
workbook.save(target_file_path)

print("处理完成！")