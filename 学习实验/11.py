import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy


def copy_sheet(template_wb, target_wb, template_sheet_name, target_sheet_name):
    # 获取模板中的"字典sheet页"
    template_sheet = template_wb[template_sheet_name]

    # 检查目标文件中是否存在相同的Sheet名称
    if target_sheet_name in target_wb.sheetnames:
        target_sheet = target_wb[target_sheet_name]
        # 清空目标sheet页内容
        for row in target_sheet.iter_rows():
            for cell in row:
                cell.value = None
    else:
        # 如果不存在则创建新的Sheet
        target_sheet = target_wb.create_sheet(title=target_sheet_name)

    # 复制模板sheet页的内容到目标sheet页
    for template_row, target_row in zip(template_sheet.iter_rows(), target_sheet.iter_rows()):
        for template_cell, target_cell in zip(template_row, target_row):
            target_cell.value = template_cell.value
            if template_cell.has_style:
                target_cell._style = copy(template_cell._style)

            # 如果需要复制其他属性如数据验证等，可以在此处添加
            # target_cell.data_type = template_cell.data_type
            # target_cell.error_value = template_cell.error_value
            # 等等...


def main():
    # 定义模板文件和目标文件路径
    template_file_path = r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\0 梳理工具\字典模板-SJ.xlsx'
    target_file_path = r'E:\工作\2.方案设计\2 数采工作\1 各基地\01 石家庄基地\2 石家庄梳理成果\04.大宗气体系统\To-Internal_设备编码清单_大宗气体_SJ.xlsx'

    # 加载模板文件
    template_wb = openpyxl.load_workbook(template_file_path, data_only=True)

    # 加载目标文件
    target_wb = openpyxl.load_workbook(target_file_path, keep_vba=True)

    # 复制字典sheet页
    copy_sheet(template_wb, target_wb, '字典-设备&仪表管理', '字典-设备&仪表管理')

    # 保存修改后的目标文件
    target_wb.save(target_file_path)


if __name__ == "__main__":
    main()